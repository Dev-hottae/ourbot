import datetime
import time

import requests
from openpyxl import load_workbook
from pytz import timezone

HOST = "https://api.binance.com"
# 과거 데이터 호출
def prev_data_request(market, limit, interval="1d"):
    endpoint = "/api/v3/klines"

    query = {
        "symbol": market,
        "interval": interval,
        "limit": limit
    }

    url = HOST + endpoint

    res = requests.get(url, params=query)
    data = res.json()

    last_data_time = datetime.datetime.fromtimestamp(data[limit-1][0] / 1000, timezone('UTC')).isoformat()
    on_time = datetime.datetime.now(timezone('UTC')).strftime('%Y-%m-%d')

    timer = 0
    while (on_time not in last_data_time) | (timer == 10):
        res = requests.get(url, params=query)
        data = res.json()
        timer += 1
        time.sleep(1)
    else:
        pass

    data_list = []

    for i in range(len(data) - 1, -1, -1):
        _time = datetime.datetime.fromtimestamp(data[i][0] / 1000).isoformat()
        open = float(data[i][1])
        high = float(data[i][2])
        low = float(data[i][3])
        close = float(data[i][4])

        one_data = {
            "candle_date_time_kst": _time,
            "opening_price": open,
            "high_price": high,
            "low_price": low,
            "trade_price": close
        }
        data_list.append(one_data)

    return data_list

data = prev_data_request("BNBUSDT", 894, interval="1d")
print(data)

wb = load_workbook('./test.xlsx')  # 기존 파일 불러옴.

sheet = wb['Sheet1']


# 파일 쓰기
sheet['A1'] = 'date'
sheet['D1'] = 'open'
sheet['E1'] = 'high'
sheet['F1'] = 'low'
sheet['G1'] = 'close'

for i in range(len(data)):
    sheet['A'+str(i+2)] = data[i]["candle_date_time_kst"]
    sheet['D' + str(i + 2)] = data[i]["opening_price"]
    sheet['E' + str(i + 2)] = data[i]["high_price"]
    sheet['F' + str(i + 2)] = data[i]["low_price"]
    sheet['G' + str(i + 2)] = data[i]["trade_price"]


# abc = [1, 2, 'abc', 4, 55, 3, 45, 5, 2, 32, 43, 3, 2, 2, 3, 5, 5]
#
# for rowindex in range(1, 10):
#     sheet.cell(row=rowindex, column=1).value = rowindex
#
#     sheet.cell(row=rowindex, column=2).value = abc[(rowindex - 1)]

wb.save('./test.xlsx')  # 파일로 다시 저장.



